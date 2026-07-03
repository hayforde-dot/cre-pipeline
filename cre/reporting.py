"""STAGE 5 — LP Reporting.
Consumes distributions + capital_accounts records and produces:
  1. An LP statement workbook: one sheet per partner (period-by-period
     contributions, distributions by waterfall tier, running capital
     position) + a summary sheet. Totals/EMx/IRR are live Excel
     formulas so the statement self-verifies on recalc.
  2. Per-partner markdown statements (email-ready text).
Placeholder-driven runs carry an explicit banner on every statement.
"""
from __future__ import annotations
import sqlite3
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

MONEY = '#,##0;(#,##0);"-"'
PCT = '0.0%'
HDR_FILL = PatternFill('solid', start_color='1F3864')
BAND = PatternFill('solid', start_color='F2F2F2')
WARN_FILL = PatternFill('solid', start_color='FFF2CC')
THIN = Border(bottom=Side(style='thin', color='BFBFBF'))


def _tier_labels(con, entity_id):
    tiers = con.execute(
        "SELECT track, tier_no, hurdle_rate, promote_pct FROM waterfall_tiers "
        "WHERE entity_id=? ORDER BY track, tier_no", (entity_id,)).fetchall()
    labels = {}
    for t in tiers:
        pre = "" if t["track"] == "all" else ("Ops " if t["track"] == "operating" else "Cap Evt ")
        if t["hurdle_rate"] is None:
            lbl = f"{pre}Tier {t['tier_no']} Residual"
        else:
            lbl = f"{pre}Tier {t['tier_no']} to {t['hurdle_rate']:.0%}"
        labels[(t["track"], t["tier_no"])] = lbl
    return labels


def _statement_rows(con, entity_id, scenario_id, partner_id):
    accts = con.execute(
        "SELECT * FROM capital_accounts WHERE entity_id=? AND scenario_id=? AND partner_id=? "
        "ORDER BY period", (entity_id, scenario_id, partner_id)).fetchall()
    dists = con.execute(
        "SELECT period, track, tier_no, amount FROM distributions "
        "WHERE entity_id=? AND scenario_id=? AND partner_id=?",
        (entity_id, scenario_id, partner_id)).fetchall()
    by_pt = {}
    for d in dists:
        by_pt.setdefault(d["period"], {})[(d["track"], d["tier_no"])] = d["amount"]
    return accts, by_pt


def write_lp_statement_workbook(con, entity_id, scenario_id, path,
                                as_of: date, period_dates: list[str],
                                placeholder_notes: list[dict] | None = None):
    ent = con.execute("SELECT * FROM partnership_entities WHERE entity_id=?",
                      (entity_id,)).fetchone()
    deal = con.execute("SELECT * FROM deals WHERE deal_id=?", (ent["deal_id"],)).fetchone()
    scen = con.execute("SELECT * FROM scenarios WHERE scenario_id=?", (scenario_id,)).fetchone()
    partners = con.execute("SELECT * FROM partners WHERE entity_id=? ORDER BY role DESC",
                           (entity_id,)).fetchall()
    labels = _tier_labels(con, entity_id)
    tier_keys = list(labels.keys())

    wb = Workbook()
    summ = wb.active
    summ.title = "Summary"

    def banner(ws, ncols):
        if placeholder_notes:
            ws.merge_cells(start_row=ws.max_row + 1, start_column=1,
                           end_row=ws.max_row + 1, end_column=ncols)
            c = ws.cell(ws.max_row, 1,
                        "PREPARED FROM PLACEHOLDER ASSUMPTIONS — ILLUSTRATIVE ONLY, "
                        "NOT ACTUAL RESULTS. See Assumptions sheet.")
            c.font = Font(name='Arial', bold=True, color='9C5700')
            c.fill = WARN_FILL

    # ---------- per-partner sheets ----------
    partner_refs = {}
    for p in partners:
        ws = wb.create_sheet(p["name"][:31])
        ws.sheet_view.showGridLines = False
        t = ws.cell(1, 1, f"{deal['name']} — {ent['name']}")
        t.font = Font(name='Arial', bold=True, size=14)
        ws.cell(2, 1, f"Distribution Statement — {p['name']} ({p['role']}, "
                      f"{p['equity_pct']:.0%} of equity)").font = Font(name='Arial', size=11)
        ws.cell(3, 1, f"Scenario: {scen['name']}   |   As of: {as_of.isoformat()}"
                ).font = Font(name='Arial', size=10, italic=True)
        banner(ws, 8 + len(tier_keys))

        hdr_row = ws.max_row + 2
        heads = (["Period", "Date", "Contributions"] + [labels[k] for k in tier_keys]
                 + ["Total Distributions", "Net Cash Flow",
                    "Cumulative Contributed", "Cumulative Distributed", "Net Position"])
        for j, h in enumerate(heads, start=1):
            c = ws.cell(hdr_row, j, h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        accts, by_pt = _statement_rows(con, entity_id, scenario_id, p["partner_id"])
        r0 = hdr_row + 1
        for i, a in enumerate(accts):
            r = r0 + i
            per = a["period"]
            ws.cell(r, 1, per)
            ws.cell(r, 2, period_dates[per] if per < len(period_dates) else "")
            ws.cell(r, 3, -a["contribution"])          # shown as outflow
            col = 4
            for k in tier_keys:
                ws.cell(r, col, by_pt.get(per, {}).get(k, 0.0))
                col += 1
            first_t, last_t = gcl(4), gcl(3 + len(tier_keys))
            ws.cell(r, col, f"=SUM({first_t}{r}:{last_t}{r})")
            ws.cell(r, col + 1, f"={gcl(col)}{r}+C{r}")
            ws.cell(r, col + 2, f"=-SUM($C${r0}:$C{r})")
            ws.cell(r, col + 3, f"=SUM({gcl(col)}${r0}:{gcl(col)}{r})")
            ws.cell(r, col + 4, f"={gcl(col+3)}{r}-{gcl(col+2)}{r}")
            for j in range(3, col + 5):
                ws.cell(r, j).number_format = MONEY
                ws.cell(r, j).font = Font(name='Arial', size=9)
                if i % 2:
                    ws.cell(r, j).fill = BAND
        rN = r0 + len(accts) - 1
        tr = rN + 1
        ws.cell(tr, 2, "Total").font = Font(name='Arial', bold=True)
        for j in range(3, 4 + len(tier_keys) + 2):
            ws.cell(tr, j, f"=SUM({gcl(j)}{r0}:{gcl(j)}{rN})")
            ws.cell(tr, j).number_format = MONEY
            ws.cell(tr, j).font = Font(name='Arial', bold=True)
            ws.cell(tr, j).border = Border(top=Side(style='double'))

        sb = tr + 2
        ncf_col = gcl(4 + len(tier_keys) + 1)
        stats = [
            ("Total Contributed", f"=-C{tr}", MONEY),
            ("Total Distributed", f"={gcl(4+len(tier_keys))}{tr}", MONEY),
            ("Profit", f"={ncf_col}{tr}", MONEY),
            ("Equity Multiple (DPI)",
             f"=SUMIF({ncf_col}{r0}:{ncf_col}{rN},\">0\")/-SUMIF({ncf_col}{r0}:{ncf_col}{rN},\"<0\")",
             '0.00"x"'),
            ("IRR (periodic)", f"=IRR({ncf_col}{r0}:{ncf_col}{rN})", PCT),
        ]
        for i, (lbl, f, fmt) in enumerate(stats):
            ws.cell(sb + i, 2, lbl).font = Font(name='Arial', bold=True, size=10)
            c = ws.cell(sb + i, 3, f)
            c.number_format = fmt
            c.font = Font(name='Arial', size=10)
        partner_refs[p["name"]] = (ws.title, sb)
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 11
        for j in range(3, 9 + len(tier_keys)):
            ws.column_dimensions[gcl(j)].width = 15

    # ---------- summary sheet ----------
    summ.sheet_view.showGridLines = False
    summ.cell(1, 1, f"{deal['name']} — {ent['name']} — Partner Summary"
              ).font = Font(name='Arial', bold=True, size=14)
    summ.cell(2, 1, f"Scenario: {scen['name']}   |   As of: {as_of.isoformat()}"
              ).font = Font(name='Arial', italic=True, size=10)
    banner(summ, 7)
    hr = summ.max_row + 2
    for j, h in enumerate(["Partner", "Role", "Equity %", "Contributed",
                           "Distributed", "Profit", "EMx", "IRR"], start=1):
        c = summ.cell(hr, j, h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = HDR_FILL
    for i, p in enumerate(partners):
        r = hr + 1 + i
        sheet, sb = partner_refs[p["name"]]
        summ.cell(r, 1, p["name"]).font = Font(name='Arial', size=10)
        summ.cell(r, 2, p["role"]).font = Font(name='Arial', size=10)
        summ.cell(r, 3, p["equity_pct"]).number_format = PCT
        for j, off in ((4, 0), (5, 1), (6, 2)):
            summ.cell(r, j, f"='{sheet}'!C{sb + off}").number_format = MONEY
        summ.cell(r, 7, f"='{sheet}'!C{sb + 3}").number_format = '0.00"x"'
        summ.cell(r, 8, f"='{sheet}'!C{sb + 4}").number_format = PCT
        for j in range(1, 9):
            summ.cell(r, j).font = Font(name='Arial', size=10)
    for col, w in zip("ABCDEFGH", (22, 8, 9, 14, 14, 14, 9, 9)):
        summ.column_dimensions[col].width = w

    # ---------- assumptions sheet ----------
    if placeholder_notes:
        aw = wb.create_sheet("Assumptions")
        aw.cell(1, 1, "PLACEHOLDER ASSUMPTIONS USED IN THIS STATEMENT"
                ).font = Font(name='Arial', bold=True, size=12)
        aw.cell(2, 1, "Every value below was assumed, not provided. Replace before "
                      "any external use.").font = Font(name='Arial', italic=True, size=10)
        for j, h in enumerate(["Assumption", "Value", "Note"], start=1):
            c = aw.cell(4, j, h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF')
            c.fill = HDR_FILL
        for i, a in enumerate(placeholder_notes):
            aw.cell(5 + i, 1, a["key"]).font = Font(name='Arial', size=10)
            aw.cell(5 + i, 2, a["value"]).font = Font(name='Arial', size=10)
            aw.cell(5 + i, 3, a.get("note") or "").font = Font(name='Arial', size=10)
        aw.column_dimensions['A'].width = 30
        aw.column_dimensions['B'].width = 14
        aw.column_dimensions['C'].width = 60
    wb.save(path)
    return path


def write_lp_statement_markdown(con, entity_id, scenario_id, partner_id,
                                period_dates, placeholder: bool) -> str:
    ent = con.execute("SELECT * FROM partnership_entities WHERE entity_id=?",
                      (entity_id,)).fetchone()
    deal = con.execute("SELECT * FROM deals WHERE deal_id=?", (ent["deal_id"],)).fetchone()
    p = con.execute("SELECT * FROM partners WHERE partner_id=?", (partner_id,)).fetchone()
    accts, by_pt = _statement_rows(con, entity_id, scenario_id, partner_id)
    labels = _tier_labels(con, entity_id)
    lines = [f"# {deal['name']} — Distribution Statement",
             f"**Partner:** {p['name']} ({p['role']}, {p['equity_pct']:.0%} of equity)  ",
             f"**Entity:** {ent['name']}", ""]
    if placeholder:
        lines += ["> **PREPARED FROM PLACEHOLDER ASSUMPTIONS — illustrative only, "
                  "not actual results.**", ""]
    heads = ["Period", "Date", "Contribution"] + [labels[k] for k in labels] + \
            ["Total Dist", "Cum Net Position"]
    lines.append("| " + " | ".join(heads) + " |")
    lines.append("|" + "---|" * len(heads))
    for a in accts:
        per = a["period"]
        tier_cells = [f"{by_pt.get(per, {}).get(k, 0):,.0f}" for k in labels]
        lines.append("| " + " | ".join(
            [str(per), period_dates[per] if per < len(period_dates) else "",
             f"({a['contribution']:,.0f})" if a["contribution"] else "-"]
            + tier_cells + [f"{a['distribution']:,.0f}", f"{a['net_position']:,.0f}"]) + " |")
    tc = sum(a["contribution"] for a in accts)
    td = sum(a["distribution"] for a in accts)
    lines += ["", f"**Total contributed:** {tc:,.0f}   **Total distributed:** {td:,.0f}   "
                  f"**Profit:** {td - tc:,.0f}   **DPI:** {td / tc:.2f}x" if tc else ""]
    return "\n".join(lines)


def write_underwriting_summary(con, deal_id, path, placeholder_notes=None):
    deal = con.execute("SELECT * FROM deals WHERE deal_id=?", (deal_id,)).fetchone()
    rows = con.execute(
        """SELECT s.name scenario, u.structure, u.metric, u.value
           FROM underwriting_results u JOIN scenarios s ON s.scenario_id=u.scenario_id
           WHERE u.deal_id=? ORDER BY s.scenario_id""", (deal_id,)).fetchall()
    combos, seen = [], set()
    for r in rows:
        k = (r["scenario"], r["structure"])
        if k not in seen:
            seen.add(k)
            combos.append(k)
    vals = {(r["scenario"], r["structure"], r["metric"]): r["value"] for r in rows}
    metrics = [
        ("purchase_price", "Purchase Price", MONEY), ("all_in_cost", "All-in Cost", MONEY),
        ("going_in_cap", "Going-in Cap Rate", PCT), ("sale_price", "Gross Sale Price", MONEY),
        ("total_debt", "Total Debt", MONEY), ("ltv", "LTV", PCT),
        ("equity", "Equity Required", MONEY), ("dscr_y1", "DSCR (Yr 1)", '0.00"x"'),
        ("dscr_min", "DSCR (Min)", '0.00"x"'), ("debt_yield_y1", "Debt Yield (Yr 1)", PCT),
        ("coc_y1", "Cash-on-Cash (Yr 1)", PCT), ("coc_avg", "Cash-on-Cash (Avg)", PCT),
        ("unlevered_irr", "Unlevered IRR", PCT), ("levered_irr", "Levered IRR", PCT),
        ("unlevered_em", "Unlevered EMx", '0.00"x"'), ("levered_em", "Levered EMx", '0.00"x"'),
        ("levered_profit", "Levered Profit", MONEY),
        ("unlevered_irr_cfo_share", "IRR from Operations", PCT),
        ("unlevered_irr_reversion_share", "IRR from Reversion", PCT),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Underwriting Summary"
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, f"{deal['name']} — Underwriting Summary").font = Font(
        name='Arial', bold=True, size=14)
    r = 2
    if placeholder_notes:
        c = ws.cell(r, 1, "PREPARED FROM PLACEHOLDER ASSUMPTIONS — ILLUSTRATIVE ONLY")
        c.font = Font(name='Arial', bold=True, color='9C5700')
        c.fill = WARN_FILL
        r += 1
    hr = r + 1
    ws.cell(hr, 1, "Metric").font = Font(name='Arial', bold=True, color='FFFFFF')
    ws.cell(hr, 1).fill = HDR_FILL
    for j, (sc, st) in enumerate(combos, start=2):
        c = ws.cell(hr, j, f"{sc}\n{st}")
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, horizontal='center')
        ws.column_dimensions[gcl(j)].width = 15
    for i, (key, lbl, fmt) in enumerate(metrics):
        rr = hr + 1 + i
        ws.cell(rr, 1, lbl).font = Font(name='Arial', size=10)
        ws.cell(rr, 1).border = THIN
        for j, (sc, st) in enumerate(combos, start=2):
            v = vals.get((sc, st, key))
            c = ws.cell(rr, j, v if v is not None else "-")
            c.number_format = fmt
            c.font = Font(name='Arial', size=10)
            c.border = THIN
    ws.column_dimensions['A'].width = 24
    loans = con.execute("SELECT * FROM loans WHERE deal_id=? ORDER BY scenario_id, structure",
                        (deal_id,)).fetchall()
    lw = wb.create_sheet("Debt Stack")
    for j, h in enumerate(["Scenario", "Structure", "Tranche", "Amount", "Rate",
                           "Term (mo)", "Amort (mo)", "IO (mo)", "Binding Constraint"], 1):
        c = lw.cell(1, j, h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = HDR_FILL
        lw.column_dimensions[gcl(j)].width = 14
    scen_names = {s["scenario_id"]: s["name"] for s in
                  con.execute("SELECT * FROM scenarios WHERE deal_id=?", (deal_id,))}
    for i, l in enumerate(loans):
        r = 2 + i
        cells = [scen_names.get(l["scenario_id"], ""), l["structure"], l["tranche"],
                 l["sized_amount"], l["rate"], l["term_months"], l["amort_months"],
                 l["io_months"], l["binding_constraint"]]
        for j, v in enumerate(cells, 1):
            c = lw.cell(r, j, v)
            c.font = Font(name='Arial', size=10)
            if j == 4:
                c.number_format = MONEY
            if j == 5:
                c.number_format = '0.00%'
    wb.save(path)
    return path


def write_valuation_workbook(con, deal_id, scenario_id, valuation: dict, path,
                             placeholder_notes=None):
    """Two-sheet workbook mirroring the firm's income-approach templates:
    Method #1 Direct Cap (reserve-exclusive NOI / cap) and Method #2 DCF
    (price from going-in cap, exit at drifted going-out cap, IRR/EM/Profit
    as live formulas)."""
    deal = con.execute("SELECT * FROM deals WHERE deal_id=?", (deal_id,)).fetchone()
    pf1 = con.execute("""SELECT * FROM proforma_lines WHERE deal_id=? AND
        scenario_id=? AND year=1""", (deal_id, scenario_id)).fetchone()
    exps = con.execute("SELECT * FROM expense_lines WHERE deal_id=?", (deal_id,)).fetchall()
    units = deal["units"] or 0
    sf = units * (deal["avg_unit_sf"] or 0)
    wb = Workbook()

    def _hdr(ws, r, texts):
        for j, t in enumerate(texts, start=1):
            c = ws.cell(r, j, t)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            c.fill = HDR_FILL
        return r + 1

    def _banner(ws, ncols):
        if placeholder_notes:
            ws.merge_cells(start_row=ws.max_row + 1, start_column=1,
                           end_row=ws.max_row + 1, end_column=ncols)
            c = ws.cell(ws.max_row, 1, "PREPARED FROM PLACEHOLDER ASSUMPTIONS — "
                                       "ILLUSTRATIVE ONLY, NOT AN APPRAISAL")
            c.font = Font(name='Arial', bold=True, color='9C5700')
            c.fill = WARN_FILL

    # ---------------- Method #1: Direct Cap ----------------
    dc = valuation["direct_cap"]
    ws = wb.active
    ws.title = "Direct Cap"
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "INCOME APPROACH — METHOD #1 DIRECT CAPITALIZATION"
            ).font = Font(name='Arial', bold=True, size=13)
    ws.cell(2, 1, f"{deal['name']}").font = Font(name='Arial', size=11)
    _banner(ws, 5)
    r = _hdr(ws, ws.max_row + 2, ["Stabilized Pro Forma (Year 1)", "Amount",
                                  "/Unit/Yr", "/SF/Mo", "Underwriting Notes"])
    rev_rows = [("Gross Potential Rent", pf1["gpr"], "rent roll x 12, grown"),
                ("Loss-to-Lease", pf1["loss_to_lease"], "% of GPR"),
                ("Concessions", pf1["concessions"], "% of GPR"),
                ("Other Income", pf1["other_income"], "grown from year-1 inputs"),
                ("Gross Revenue", pf1["gross_revenue"], ""),
                ("General Vacancy & Credit Loss", pf1["general_vacancy"],
                 "% of Gross Revenue"),
                ("Effective Gross Revenue", pf1["egr"], "")]
    first = r
    for name, amt, note in rev_rows:
        ws.cell(r, 1, name)
        ws.cell(r, 2, amt).number_format = MONEY
        ws.cell(r, 3, amt / units if units else None).number_format = MONEY
        ws.cell(r, 4, amt / sf / 12 if sf else None).number_format = '0.00'
        ws.cell(r, 5, note).font = Font(name='Arial', italic=True, size=9)
        for j in (1, 2, 3):
            ws.cell(r, j).font = Font(name='Arial', size=10,
                                      bold=name in ("Effective Gross Revenue",
                                                    "Gross Revenue"))
        r += 1
    r += 1
    r = _hdr(ws, r, ["Operating Expenses (Year 1)", "Amount", "/Unit/Yr", "", "Notes"])
    opex_total = 0.0
    for e in exps:
        if e["pct_of_egr"] is not None:
            amt = e["pct_of_egr"] * pf1["egr"]
            note = f"{e['pct_of_egr']:.0%} of EGR"
        else:
            amt = e["year1_amount"] or 0.0
            note = f"year-1 input, grows {e['growth_rate'] or 0:.1%}/yr"
        opex_total += amt
        ws.cell(r, 1, e["name"]).font = Font(name='Arial', size=10)
        ws.cell(r, 2, amt).number_format = MONEY
        ws.cell(r, 3, amt / units if units else None).number_format = MONEY
        ws.cell(r, 5, note).font = Font(name='Arial', italic=True, size=9)
        r += 1
    for label, amt, bold in [("Total Operating Expenses", opex_total, True),
                             ("NET OPERATING INCOME", dc["noi"], True),
                             ("Capital Reserves (below NOI)", dc["reserves"], False),
                             ("Cash Flow from Operations", dc["cfo"], True)]:
        ws.cell(r, 1, label).font = Font(name='Arial', bold=bold, size=10)
        ws.cell(r, 2, amt).number_format = MONEY
        ws.cell(r, 2).font = Font(name='Arial', bold=bold, size=10)
        r += 1
    r += 1
    r = _hdr(ws, r, ["Valuation", "", "", "", ""])
    val_rows = [("Market Cap Rate", dc["cap_rate"], PCT),
                ("Indicated Value  (NOI ÷ Cap)", dc["value"], MONEY),
                ("Value / Unit", dc["value_per_unit"], MONEY),
                ("Value / SF", dc["value_per_sf"], '0.00'),
                ("Purchase Price", dc["purchase_price"], MONEY),
                ("Indicated Value vs Price", dc["premium_to_purchase"], '+0.0%;-0.0%')]
    for name, v, fmt in val_rows:
        ws.cell(r, 1, name).font = Font(name='Arial', size=10, bold="Indicated" in name)
        c = ws.cell(r, 2, v)
        c.number_format = fmt
        c.font = Font(name='Arial', size=10, bold="Indicated" in name)
        r += 1
    r += 1
    r = _hdr(ws, r, ["Cap Rate Sensitivity", "Value", "", "", ""])
    for cap, v in dc["sensitivity"]:
        ws.cell(r, 1, cap).number_format = PCT
        ws.cell(r, 2, v).number_format = MONEY
        for j in (1, 2):
            ws.cell(r, j).font = Font(name='Arial', size=10)
        r += 1
    for col, w in zip("ABCDE", (34, 15, 12, 10, 40)):
        ws.column_dimensions[col].width = w

    # ---------------- Method #2: DCF ----------------
    d = valuation["dcf"]
    wd = wb.create_sheet("DCF")
    wd.sheet_view.showGridLines = False
    wd.cell(1, 1, "INCOME APPROACH — METHOD #2 DISCOUNTED CASH FLOW"
            ).font = Font(name='Arial', bold=True, size=13)
    wd.cell(2, 1, f"{deal['name']}").font = Font(name='Arial', size=11)
    _banner(wd, 4)
    r = wd.max_row + 2
    assum = [("Going-in Cap Rate", d["going_in_cap"], PCT),
             ("Going-out Cap Rate (going-in + 5bps × hold yrs)",
              d["going_out_cap"], PCT),
             ("Indicated Price  (Year-1 NOI ÷ Going-in Cap)", d["indicated_price"], MONEY),
             ("All-in Cost (incl. closing)", d["all_in_cost"], MONEY),
             ("Gross Sale Price (Forward NOI ÷ Going-out Cap)", d["sale_price"], MONEY),
             ("Net Sale Proceeds", d["sale_proceeds"], MONEY),
             ("Purchase Price (actual)", d["purchase_price"], MONEY),
             ("Indicated Price vs Purchase Price", d["premium_to_purchase"],
              '+0.0%;-0.0%')]
    for name, v, fmt in assum:
        wd.cell(r, 1, name).font = Font(name='Arial', size=10, bold="Indicated" in name)
        c = wd.cell(r, 2, v)
        c.number_format = fmt
        c.font = Font(name='Arial', size=10, bold="Indicated" in name)
        r += 1
    r += 1
    r = _hdr(wd, r, ["Year", "Unlevered Cash Flow", "", ""])
    r0 = r
    for p, cf in enumerate(d["cash_flows"]):
        wd.cell(r, 1, p).font = Font(name='Arial', size=10)
        wd.cell(r, 2, cf).number_format = MONEY
        wd.cell(r, 2).font = Font(name='Arial', size=10)
        r += 1
    rN = r - 1
    r += 1
    ret = [("Unlevered IRR", f"=IRR(B{r0}:B{rN})", PCT),
           ("Equity Multiple",
            f'=SUMIF(B{r0}:B{rN},">0")/-SUMIF(B{r0}:B{rN},"<0")', '0.00"x"'),
           ("Profit", f"=SUM(B{r0}:B{rN})", MONEY),
           ("Supplementary: IRR at actual purchase price",
            d["irr_at_purchase_price"], PCT)]
    for name, v, fmt in ret:
        wd.cell(r, 1, name).font = Font(name='Arial', bold=True, size=10)
        c = wd.cell(r, 2, v)
        c.number_format = fmt
        c.font = Font(name='Arial', size=10)
        r += 1
    wd.column_dimensions['A'].width = 44
    wd.column_dimensions['B'].width = 17
    wb.save(path)
    return path
