"""Stage 4 + Stage 2 verification against the firm's actual workbooks.
These tests read the uploaded Excel files and assert the Python engines
reproduce their numbers — per-period vectors, not just totals.
"""
import math
from datetime import date
from pathlib import Path
import openpyxl
import pytest

from cre.db import connect
from cre.capital import create_entity
from cre.waterfall import run_irr_hurdle, run_american_dual_track, partner_metrics
from cre.underwriting import size_senior, annual_debt_service, loan_payoff
from cre.finance import xirr, irr

UPLOADS = Path("/mnt/user-data/uploads")
WF = UPLOADS / "Waterfalls-Completed_2.xlsx"
AM = UPLOADS / "American-Style-Equity-Waterfall.xlsx"


def _row(ws, r, c0, c1):
    return [ws.cell(r, c).value or 0.0 for c in range(c0, c1 + 1)]


@pytest.fixture(scope="module")
def gp_lp_sheet():
    wb = openpyxl.load_workbook(WF, data_only=True)
    return wb["GP-LP Returns"]


@pytest.fixture(scope="module")
def deal_con():
    con = connect()
    con.execute("""INSERT INTO deals (deal_id, name, hold_period_years,
                   periods_per_year, purchase_price, exit_cap_rate)
                   VALUES (1,'Frost and Main',10,12,0,0.05)""")
    con.execute("INSERT INTO scenarios (scenario_id, deal_id, name) VALUES (1,1,'base')")
    con.commit()
    return con


def test_gp_lp_waterfall_replicates_workbook(gp_lp_sheet, deal_con):
    ws = gp_lp_sheet
    # columns F..DV = 6..126 hold the monthly periods
    c0, c1 = 6, 126
    net_cf = _row(ws, 34, c0, c1)
    ent = create_entity(deal_con, 1, "GP-LP", "irr_hurdle",
                        partners=[{"name": "LP", "role": "LP", "equity_pct": 0.9},
                                  {"name": "GP", "role": "GP", "equity_pct": 0.1}],
                        tiers=[{"tier_no": 1, "hurdle_rate": 0.08, "promote_pct": 0.0},
                               {"tier_no": 2, "hurdle_rate": 0.11, "promote_pct": 0.2},
                               {"tier_no": 3, "hurdle_rate": 0.14, "promote_pct": 0.3},
                               {"tier_no": 4, "hurdle_rate": None, "promote_pct": 0.4}])
    res = run_irr_hurdle(deal_con, ent["entity_id"], net_cf, periods_per_year=12)
    lp_id = ent["partner_ids"]["LP"]
    gp_id = ent["partner_ids"]["GP"]

    # per-period LP distributions = rows 47 + 62 + 77 + 84
    lp_sheet = [a + b + c + d for a, b, c, d in zip(
        _row(ws, 47, c0, c1), _row(ws, 62, c0, c1),
        _row(ws, 77, c0, c1), _row(ws, 84, c0, c1))]
    gp_sheet = [a + b + c + d for a, b, c, d in zip(
        _row(ws, 48, c0, c1), _row(ws, 63, c0, c1),
        _row(ws, 78, c0, c1), _row(ws, 85, c0, c1))]
    for p in range(len(net_cf)):
        assert res.distributions[lp_id][p] == pytest.approx(lp_sheet[p], abs=0.01), f"LP period {p}"
        assert res.distributions[gp_id][p] == pytest.approx(gp_sheet[p], abs=0.01), f"GP period {p}"

    # hurdle-1 ending balances, full vector (row 44)
    bal_sheet = _row(ws, 44, c0, c1)
    for p, (mine, theirs) in enumerate(zip(res.hurdle_accounts[("all", 1)], bal_sheet)):
        assert mine == pytest.approx(theirs, abs=0.01), f"tier1 balance period {p}"

    # headline summary block (D19:D30)
    assert sum(res.distributions[lp_id]) == pytest.approx(ws["D19"].value, abs=0.01)
    assert sum(res.contributions[lp_id]) == pytest.approx(ws["D20"].value, abs=0.01)
    assert sum(res.distributions[gp_id]) == pytest.approx(ws["D26"].value, abs=0.01)
    assert sum(res.contributions[gp_id]) == pytest.approx(ws["D27"].value, abs=0.01)

    dates = [ws.cell(15, c).value.date() for c in range(c0, c1 + 1)]
    assert xirr(res.net_cf(lp_id), dates) == pytest.approx(ws["D22"].value, abs=2e-4)
    assert xirr(res.net_cf(gp_id), dates) == pytest.approx(ws["D29"].value, abs=2e-4)


def test_co_gp_waterfall_replicates_workbook(gp_lp_sheet, deal_con):
    wb = openpyxl.load_workbook(WF, data_only=True)
    ws = wb["Co-GP Returns"]
    c0, c1 = 6, 126
    net_cf = _row(ws, 34, c0, c1)   # = GP stream from layer 1, fee 0 in file
    ent = create_entity(deal_con, 1, "Co-GP", "irr_hurdle",
                        partners=[{"name": "Investors", "role": "LP", "equity_pct": 0.95},
                                  {"name": "Sponsor", "role": "GP", "equity_pct": 0.05}],
                        tiers=[{"tier_no": 1, "hurdle_rate": 0.15, "promote_pct": 0.0},
                               {"tier_no": 2, "hurdle_rate": None, "promote_pct": 0.4}])
    res = run_irr_hurdle(deal_con, ent["entity_id"], net_cf, periods_per_year=12)
    inv, spo = ent["partner_ids"]["Investors"], ent["partner_ids"]["Sponsor"]
    assert sum(res.distributions[inv]) == pytest.approx(ws["D19"].value, abs=0.5)
    assert sum(res.distributions[spo]) == pytest.approx(ws["D26"].value, abs=0.5)
    dates = [ws.cell(15, c).value.date() for c in range(c0, c1 + 1)]
    assert xirr(res.net_cf(inv), dates) == pytest.approx(ws["D22"].value, abs=3e-4)
    assert xirr(res.net_cf(spo), dates) == pytest.approx(ws["D29"].value, abs=3e-4)


def test_american_waterfall_replicates_workbook(deal_con):
    wb = openpyxl.load_workbook(AM, data_only=True)
    ws = wb["Completed"]
    c0, c1 = 6, 16   # F..P = periods 0..10
    invest = _row(ws, 30, c0, c1)
    ops = _row(ws, 31, c0, c1)
    rev = _row(ws, 32, c0, c1)
    con = connect()
    con.execute("""INSERT INTO deals (deal_id, name, hold_period_years, periods_per_year,
                   purchase_price, exit_cap_rate) VALUES (1,'American',10,1,0,0.05)""")
    ent = create_entity(con, 1, "Partnership", "american_dual_track",
                        partners=[{"name": "LP", "role": "LP", "equity_pct": 0.9},
                                  {"name": "GP", "role": "GP", "equity_pct": 0.1}],
                        tiers=[{"track": "operating", "tier_no": 1, "hurdle_rate": 0.06, "promote_pct": 0.0},
                               {"track": "operating", "tier_no": 2, "hurdle_rate": None, "promote_pct": 0.15},
                               {"track": "capital_event", "tier_no": 1, "hurdle_rate": 0.06, "promote_pct": 0.0},
                               {"track": "capital_event", "tier_no": 2, "hurdle_rate": 0.12, "promote_pct": 0.15},
                               {"track": "capital_event", "tier_no": 3, "hurdle_rate": None, "promote_pct": 0.45}])
    res = run_american_dual_track(con, ent["entity_id"], ops, rev,
                                  funding_cf=[-x for x in invest])
    lp_id, gp_id = ent["partner_ids"]["LP"], ent["partner_ids"]["GP"]
    lp_sheet, gp_sheet = _row(ws, 21, c0, c1), _row(ws, 26, c0, c1)
    for p in range(len(ops)):
        assert res.distributions[lp_id][p] == pytest.approx(lp_sheet[p], abs=0.01), f"LP p{p}"
        assert res.distributions[gp_id][p] == pytest.approx(gp_sheet[p], abs=0.01), f"GP p{p}"
    # capital-event tier 1 account, full vector (row 61)
    bal = _row(ws, 61, c0, c1)
    for p, (mine, theirs) in enumerate(zip(res.hurdle_accounts[("capital_event", 1)], bal)):
        assert mine == pytest.approx(theirs, abs=0.01), f"event t1 balance p{p}"
    assert irr(res.net_cf(lp_id)) == pytest.approx(ws["E22"].value, abs=1e-5)
    assert irr(res.net_cf(gp_id)) == pytest.approx(ws["E27"].value, abs=1e-5)


def test_loan_sizing_replicates_financing_workbook():
    wb = openpyxl.load_workbook(UPLOADS / "Levered_Discounted_Cash_Flow_-_Financing.xlsx",
                                data_only=True)
    ws = wb["Levered DCF"]
    price, noi1 = ws["C9"].value, ws["I83"].value
    s = size_senior(price, noi1, rate=0.0525, amort_months=360,
                    max_ltv=0.65, min_dscr=1.30, min_dy=0.08)
    assert s["by_ltv"] == pytest.approx(ws["H18"].value, abs=0.01)
    assert s["by_dscr"] == pytest.approx(ws["H19"].value, abs=0.01)
    assert s["by_dy"] == pytest.approx(ws["H20"].value, abs=0.01)
    assert s["amount"] == pytest.approx(ws["H21"].value, abs=0.01)
    assert s["binding"] == "DSCR"
    loan = s["amount"]
    assert annual_debt_service(loan, 0.0525, 360, 36, 1) == pytest.approx(ws["I89"].value, abs=0.01)
    assert annual_debt_service(loan, 0.0525, 360, 36, 4) == pytest.approx(ws["L89"].value, abs=0.01)
    assert loan_payoff(loan, 0.0525, 360, 36, 120) == pytest.approx(ws["H12"].value, abs=0.01)
