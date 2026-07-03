"""API contract tests — what the Lovable frontend depends on."""
import copy
import io
import openpyxl
import pytest
from fastapi.testclient import TestClient
from api import app
from cre.intake import MELADON_PAYLOAD

client = TestClient(app)


def test_health_and_sample():
    assert client.get("/api/health").json() == {"status": "ok"}
    s = client.get("/api/sample").json()
    assert s["deal"]["name"] == "Meladon Haymarket"


def test_run_and_download_round_trip():
    r = client.post("/api/run", json=MELADON_PAYLOAD)
    assert r.status_code == 200, r.text
    body = r.json()
    wf = body["report"]["stages"]["4_waterfall"]["base"]
    assert wf["Institutional LP"]["irr"] == pytest.approx(0.10859, abs=1e-4)
    assert wf["Family Office LP"]["irr"] == pytest.approx(
        wf["Institutional LP"]["irr"], abs=1e-9)
    assert set(body["files"]) == {"lp_statements", "cogp_statements",
                                  "underwriting_summary", "valuation"}
    for f in body["files"].values():
        d = client.get(f["url"])
        assert d.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(d.content))   # valid xlsx
        assert wb.sheetnames
    # LP statement has a sheet per partner
    d = client.get(body["files"]["lp_statements"]["url"])
    wb = openpyxl.load_workbook(io.BytesIO(d.content))
    assert {"Institutional LP", "Family Office LP", "General Partner"} <= set(wb.sheetnames)


def test_no_mezz_no_cogp_still_runs():
    p = copy.deepcopy(MELADON_PAYLOAD)
    p["mezz_loan"] = None
    p["cogp"] = None
    p["is_placeholder"] = False
    r = client.post("/api/run", json=p)
    assert r.status_code == 200, r.text
    assert set(r.json()["files"]) == {"lp_statements", "underwriting_summary",
                                      "valuation"}
    assert "4_waterfall_cogp" not in r.json()["report"]["stages"]


def test_validation_errors_are_422_with_messages():
    p = copy.deepcopy(MELADON_PAYLOAD)
    p["partnership"]["partners"][0]["equity_pct"] = 0.5        # sums to 0.87
    r = client.post("/api/run", json=p)
    assert r.status_code == 422
    assert any("sum" in e for e in r.json()["detail"])

    p2 = copy.deepcopy(MELADON_PAYLOAD)
    p2["partnership"]["tiers"][-1]["hurdle_rate"] = 0.20       # no residual tier
    r2 = client.post("/api/run", json=p2)
    assert r2.status_code == 422
    assert any("residual" in e for e in r2.json()["detail"])

    r3 = client.get("/api/download/nope/nothing.xlsx")
    assert r3.status_code == 404
